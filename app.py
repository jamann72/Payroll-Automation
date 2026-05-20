"""
Payroll Reconciliation App  –  Stonelink Property Management
SOP v2.0  |  AI-Tool Enhanced Edition  (Updated 2026-04-27)

Outputs a two-tab Excel workbook:
  Tab 1 – Weekly Recap  (Planet Synergy PM Report, §5.4–5.7)
           One row per Property Meld CSV entry; hours from TSheets.
           15 Circle St entries segregated in a separate section.
  Tab 2 – Payroll Hours  (§5.11)
           Per-employee: Last Name | First Name | Regular | Vacation |
           Holiday | Sick/Flex | Overtime | Total

Key SOP v2.0 rules enforced:
  §4.1  Property Meld CSV is PRIMARY; TSheets for hours only (F/G/H)
  §4.3  7-column CSV validation before processing
  §5.6  Protected trades (Diagnostic/Inspection/Mitigation/Quote) never NB
  §5.6  RC (Resident Charge) teal highlight
  §5.8  5 approved note conventions; PS: notes highlighted yellow
  §5.9  All 8 flag checks (Duplicate, High/Low Hour, <0.10hr, Overlap,
          No Meld, 15 Circle St, Non-English, Manual Review)
  §5.9  46-category per-task hour thresholds (18,489 jobs, Apr24–Mar26)
  §7.3  Unmatched TSheets shifts appear as 'No Meld' rows
"""

import io
import math
import re
from collections import defaultdict

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

MELD_RE = re.compile(r'#?(T[A-Z0-9]{5,10})(?![A-Z0-9])')

# Keywords that make a shift entry non-billable (§5.5 / §5.6)
NON_BILLABLE_KW = [
    "non-billable", "not billable", "office", "meeting", "dump", "vertedero",
    "ofisina", "llaves", "keys", "paperwork", "estimate",
    "entrega", "reunión", "reunion",
    "these hours are not being charged",
    "mattress pick up at the office",
    "drop off", "pick up at the office",
    "clock-out",
    "parking", "car cleaning",
]

# §5.6 — Protected categories: NEVER change to NB without explicit manager auth
PROTECTED_TRADE_KW = [
    "diagnostic", "diagnóstico", "diagnostico",
    "inspection", "inspección", "inspeccion",
    "mitigation", "mitigación", "mitigacion",
    "quote",
]

# §4.3 — All 7 required Property Meld CSV columns
REQUIRED_CSV_COLS = {"Agent", "Meld", "Unit", "Title", "Description",
                     "Check-In Hours", "Address Line 1"}

# §5.9 — 15 Circle St (Rumford, RI) flagged for separate independent billing
_FIFTEEN_CIRCLE_RE = re.compile(r'15\s+circle\s+st', re.IGNORECASE)

# §5.6 — RC (Resident Charge) detection in TSheets notes
_RC_RE = re.compile(
    r'\bRC\b|\bresident[\s-]?charge\b|\bcharge\s+to\s+resident\b',
    re.IGNORECASE
)

# §5.8 — PS: (Planet Synergy instructions) — must be highlighted yellow
_PS_NOTE_RE = re.compile(r'^PS\s*:', re.IGNORECASE)

# §5.9 / Methodology — Turnover rows skip threshold checks (handled separately)
_TURNOVER_RE = re.compile(r'\bturnover\b|\bturn[\s-]?over\b', re.IGNORECASE)

# Exact SOP-required phrase for flat billing (§5.5 — zero variation allowed)
FLAT_BILL_PHRASE = "not billable- to be flat bill"

# ── Excel fill colors ──────────────────────────────────────────────────────
DARK_BLUE     = PatternFill("solid", fgColor="1F4E79")
MED_BLUE      = PatternFill("solid", fgColor="2E75B6")
LIGHT_BLUE    = PatternFill("solid", fgColor="D6E4F0")
YELLOW_FILL   = PatternFill("solid", fgColor="FFE699")   # PS: Planet Synergy notes
AMBER_FILL    = PatternFill("solid", fgColor="FFC000")   # Paying > Billable
RED_FILL      = PatternFill("solid", fgColor="FF9999")   # No match / No Meld
ORANGE_FILL   = PatternFill("solid", fgColor="FFB347")   # NB missing note (error)
GREEN_FILL    = PatternFill("solid", fgColor="C6EFCE")   # Subtotal row
PURPLE_FILL   = PatternFill("solid", fgColor="D9B3FF")   # Flat-bill entry
TEAL_FILL     = PatternFill("solid", fgColor="80CBC4")   # RC – Resident Charge
LAVENDER_FILL = PatternFill("solid", fgColor="CE93D8")   # 15 Circle St
UNDER_FILL    = PatternFill("solid", fgColor="F1F8E9")   # Threshold UNDER (< p10)
OVER_FILL     = PatternFill("solid", fgColor="FBE9E7")   # Threshold OVER  (≥ p90)
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=10)
BOLD_FONT  = Font(bold=True, size=10)
BASE_FONT  = Font(size=10)

_THIN = Side(border_style="thin", color="AAAAAA")
BOX   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

OT_THRESHOLD = 40.0   # hours before overtime kicks in (§5.11)


# ─────────────────────────────────────────────────────────────────────────────
# TASK THRESHOLDS  (§5.9 / §7.3)
# Source: Stonelink Property Meld, Apr 2024 – Mar 2026, 18,489 classified jobs
# under_p10: flag if actual < this  |  over_p90: flag if actual >= this
# Office/Admin has over_p90=0.0 — handled by NB detection, not threshold engine
# ─────────────────────────────────────────────────────────────────────────────

def _build_threshold_table() -> list[dict]:
    raw = [
        # (category, under_p10, over_p90, median, confidence, pattern)
        ("Trash / Dump Run / Landfill",                       0.50, 3.00, 1.36, "HIGH",
         r"\btrash\b|\bdump\b|\blandfill\b|\bdebris\b|\bgarbage\b|\brubbish\b"),
        ("General Inspection / Walkthrough",                  0.83, 3.00, 1.87, "HIGH",
         r"\binspection\b|\bwalk[\s-]?through\b|\bfinal inspect\b|\bquality check\b"
         r"|\bqc\b|\bpre[\s-]?lease\b|\bpunchlist\b|\bpunch list\b"),
        ("Office / Admin / Non-Field",                        None, 0.00, 1.73, "HIGH",
         r"\boffice\b|\bmeeting\b|\btraining\b|\bpaperwork\b|\bdocument\b"
         r"|\borientation\b|\bwelcome letter\b|\bwater meter\b|\bpicture[s]? for\b"),
        ("Fire Panel / Fire Inspection",                      0.50, 2.00, 1.28, "HIGH",
         r"\bfire panel\b|\bfire inspect\b|\bfire alarm\b|\bfire marshal\b"
         r"|\bsmoke detector\b|\bcarbon monoxide\b|\bfire extinguisher\b"
         r"|\bfire ext\b|\bcombo detector\b|\bdetector\b"),
        ("Lead Inspection / Pretreat",                        1.22, 8.50, 4.25, "HIGH",
         r"\blead\b|\bpretreat\b|\bpretratamiento\b|\bxrf\b|\bconformance\b|\bencapsul"),
        ("Door Repair / Replacement",                         0.88, 5.00, 2.26, "HIGH",
         r"\bdoor\b|\bpuerta\b"),
        ("Plumbing - Sink / Faucet",                          1.03, 4.58, 2.17, "HIGH",
         r"\bsink\b|\bfaucet\b|\bspigot\b|\bgrifo\b|\bfregadero\b|\blavamanos\b"),
        ("Plumbing - Leak (Water/Pipe)",                      0.98, 4.09, 1.93, "HIGH",
         r"\bleak\b|\bleaking\b|\bgotera\b|\bfuga\b|\bpipe burst\b|\bbroken pipe\b"
         r"|\bwater damage\b|\bceiling leak\b|\bwall leak\b|\bshut off valve\b"),
        ("HVAC - No Heat / Heating",                          0.51, 2.82, 1.48, "HIGH",
         r"\bno heat\b|\bheating\b|\bfurnace\b|\bboiler\b|\bradiator\b|\bbaseboard\b"
         r"|\bcalefacci[oó]n\b|\bcalentador\b|\bsin calor\b|\bheat check\b"),
        ("Plumbing - Shower / Tub",                           1.16, 8.05, 2.43, "HIGH",
         r"\bshower\b|\btub\b|\bbathtub\b|\bba[nñ]era\b|\bducha\b"),
        ("Lock / Lock-out / Key",                             0.75, 3.11, 1.62, "HIGH",
         r"\block\b|\block-?out\b|\bkey\b|\brekey\b|\bcerradura\b|\bllave\b"
         r"|\bgarage door lock\b"),
        ("Plumbing - Toilet",                                 1.09, 5.00, 2.08, "HIGH",
         r"\btoilet\b|\bwc\b|\binodoro\b|\bcomod[ea]\b|\bflapper\b|\bwax ring\b"),
        ("Access / Let-in / Tenant Visit",                    0.70, 3.00, 1.66, "HIGH",
         r"\baccess\b|\blet[\s-]?in\b|\bresident check\b|\btenant visit\b"
         r"|\bwalk[\s-]?in\b"),
        ("Generic Billable Repair",                           0.95, 4.00, 3.90, "HIGH",
         r"\bbillable\b|\bnon[\s-]?billable\b|\binvoic"),
        ("Electrical - Lighting / Fixture",                   0.92, 3.00, 1.78, "HIGH",
         r"\blight\b|\blamp\b|\bbulb\b|\bfixture\b|\bl[aá]mpara\b|\bbombill[oa]\b"),
        ("Window Repair / Replacement",                       1.00, 5.00, 2.22, "HIGH",
         r"\bwindow\b|\bventana\b|\bbroken glass\b|\bblind\b|\bscreen\b"),
        ("Plumbing - No Hot Water / Water Heater",            0.80, 2.54, 1.54, "HIGH",
         r"\bno hot water\b|\bhot water\b|\bwater heater\b|\btankless\b"
         r"|\bboiler.*(hot water|domestic)"),
        ("Electrical - No Power / Outage",                    0.79, 2.99, 1.67, "HIGH",
         r"\bno power\b|\bpower outage\b|\bbreaker\b|\belectrical panel\b"
         r"|\bsin luz\b|\bsin electricidad\b|\bgfci\b"),
        ("Flooring",                                          1.20, 8.00, 2.78, "HIGH",
         r"\bflooring\b|\bfloor\b|\bvinyl\b|\blaminate\b|\btile\b|\bpiso\b|\bcarpet\b"),
        ("Ceiling Repair / Damage",                           1.29, 8.02, 2.67, "HIGH",
         r"\bceiling\b|\btecho\b|\bcielo raso\b"),
        ("Plumbing - Drain / Clog",                           0.96, 4.00, 2.06, "HIGH",
         r"\bdrain\b|\bclog\b|\bsnake\b|\bbacked up\b|\bbackup\b"
         r"|\bdesag[uü]e\b|\btapa[oa]"),
        ("Materials / Home Depot / Pickup",                   1.17, 2.00, 2.50, "HIGH",
         r"\bhome depot\b|\bmaterials\b|\bsupply pickup\b|\bmaterial pickup\b"
         r"|\blowe'?s\b"),
        ("Basement / Oil Removal / Water Removal",            0.82, 3.22, 1.62, "HIGH",
         r"\bbasement\b|\boil removal\b|\bsump\b|\bs[oó]tano\b"),
        ("Paint / Painting",                                  1.54, 8.00, 3.83, "HIGH",
         r"\bpaint\b|\bpintura\b|\bpintar\b"),
        ("Cleanout",                                          0.39, 3.17, 1.25, "HIGH",
         r"\bcleanout\b|\bclean[\s-]?out\b|\bdumpster clean\b|\bexterior clean\b"),
        ("Wall Repair / Drywall / Hole",                      1.43, 8.30, 3.12, "HIGH",
         r"\bwall repair\b|\bdrywall\b|\bhole in (the )?wall\b|\bsheetrock\b"
         r"|\bspackle\b|\bpatch up\b|\bcompound\b"),
        ("Plumbing - General / Other",                        0.91, 3.54, 1.54, "HIGH",
         r"\bplumbing\b|\bplumber\b|\bplomer|\bgas smell\b|\bgas leak\b"),
        ("Pest Control",                                      0.78, 2.84, 1.62, "HIGH",
         r"\bpest\b|\brodent\b|\bmice\b|\bbed bug\b|\bcockroach\b|\broach\b"
         r"|\bplaga\b|\bratones\b"),
        ("Exterior Repair / Fence / Stairs / Porch / Pole",  1.01, 8.00, 2.16, "HIGH",
         r"\bfence\b|\bporch\b|\bstair\b|\bsiding\b|\brailing\b|\bpole\b"
         r"|\bparking post\b|\bdriveway\b|\bunregistered car\b"),
        ("Electrical - Outlet / Switch / Wiring",             1.00, 4.17, 2.20, "HIGH",
         r"\boutlet\b|\bswitch\b|\bwiring\b|\benchufe\b|\binterruptor\b"),
        ("Snow / Ice / Salt",                                 0.29, 2.23, 0.95, "HIGH",
         r"\bsnow\b|\bice\b.*\b(remov|melt|dam|salt)\b|\bnieve\b|\bplowing\b"
         r"|\bsalt[\s-]?down\b|\bsalting\b|\bsalt\b"),
        ("Electrical - General",                              0.89, 3.37, 1.86, "HIGH",
         r"\belectric\b|\bel[eé]ctric"),
        ("Space Heater Delivery",                             0.35, 2.39, 1.13, "HIGH",
         r"\bspace heater\b"),
        ("Appliance - Washer / Dryer",                        0.81, 2.59, 1.33, "HIGH",
         r"\bwasher\b|\bdryer\b|\blaundry\b|\blavadora\b|\bsecadora\b"),
        ("Roof / Roof Leak",                                  1.03, 8.25, 2.49, "HIGH",
         r"\broof\b|\bgutter\b|\bdownspout\b|\btecho exterior\b"),
        ("Appliance - Stove / Oven",                          0.85, 3.33, 1.66, "HIGH",
         r"\bstove\b|\boven\b|\brange\b|\bcocina\b|\bestufa\b|\bhorno\b"),
        ("HVAC - Thermostat",                                 0.82, 2.78, 1.52, "MEDIUM",
         r"\bthermostat\b|\btermostato\b"),
        ("Multiple Issues / General Repair",                  1.11, 4.00, 2.04, "MEDIUM",
         r"\bmultiple issue\b|\bmultiple repair\b|\bgeneral repair\b|\brepairs?\b"
         r"|\bvarios\b|\bm[uú]ltiples?\b"),
        ("Appliance - Refrigerator",                          1.01, 3.00, 1.88, "MEDIUM",
         r"\brefrigerator\b|\bfridge\b|\bnevera\b|\brefrigeradora\b"),
        ("Bathroom (general / fixtures)",                     0.87, 3.93, 2.08, "MEDIUM",
         r"\bbathroom\b|\btowel rack\b|\bba[nñ]o\b"),
        ("Appliance - Dishwasher",                            1.44, 2.92, 2.31, "LOW",
         r"\bdishwasher\b|\blavavajillas\b"),
        ("Landscaping / Grounds",                             0.54, 3.50, 1.79, "LOW",
         r"\blandscap\b|\blawn\b|\bgrass\b|\bmulch\b|\bweed\b|\bjardin|\bpool\b"),
        ("Vehicle / Truck / Driving",                         0.69, 2.22, 1.06, "LOW",
         r"\bvehicle\b|\btruck repair\b|\bdriving\b"),
        ("Appliance - General",                               0.96, 2.51, 1.47, "INSUFFICIENT",
         r"\bappliance\b|\belectrodom"),
        ("HVAC - AC / Air Conditioning",                      0.76, 2.96, 1.91, "INSUFFICIENT",
         r"\bair conditi\b|\bac unit\b|\baire acondic"),
        ("Housing / Section 8 Inspection",                    0.85, 3.31, 1.65, "INSUFFICIENT",
         r"\bsection 8\b|\bhousing inspect\b|\brihousing\b|\bhud inspect\b|\bnspire\b"),
    ]
    table = []
    for category, under, over, median, confidence, pattern_str in raw:
        try:
            compiled = re.compile(pattern_str, re.IGNORECASE | re.DOTALL)
        except re.error:
            compiled = None
        table.append({"category": category, "under_p10": under, "over_p90": over,
                      "median": median, "confidence": confidence, "pattern": compiled})
    return table


TASK_THRESHOLDS = _build_threshold_table()

_FALLBACK_THRESHOLD = {
    "category": "General (no category match)",
    "under_p10": 0.50, "over_p90": 8.0, "median": 2.0,
    "confidence": "FALLBACK", "pattern": None,
}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def extract_meld_ids(text: str) -> list[str]:
    return list(dict.fromkeys(MELD_RE.findall(text.upper())))


def is_non_billable(notes: str) -> bool:
    low = notes.lower()
    return any(kw in low for kw in NON_BILLABLE_KW)


def is_flat_bill(notes: str) -> bool:
    return FLAT_BILL_PHRASE in notes.lower()


def is_rc(notes: str) -> bool:
    """§5.6 — Detect Resident Charge (RC) entries."""
    return bool(_RC_RE.search(notes or ""))


def _is_15_circle(address: str) -> bool:
    """§5.9 — 15 Circle St (Rumford, RI) requires separate billing."""
    return bool(_FIFTEEN_CIRCLE_RE.search(address or ""))


def is_protected_trade(trade: str, description: str = "") -> bool:
    """§5.6 — Diagnostics/Inspections/Mitigation/Quotes must never be NB without manager auth."""
    text = f"{trade} {description}".lower()
    return any(kw in text for kw in PROTECTED_TRADE_KW)


def is_turnover(trade: str, description: str = "", notes: str = "") -> bool:
    """Methodology — Turnover rows are handled by turnover-specific logic; skip thresholds."""
    text = f"{trade} {description} {notes}"
    return bool(_TURNOVER_RE.search(text))


def _norm_name(n: str) -> str:
    """Lowercase + strip punctuation + remove suffixes for name matching."""
    n = n.lower().strip()
    n = re.sub(r'[.\-]', ' ', n)
    n = re.sub(r'\b(jr|sr|ii|iii)\b', '', n)
    return re.sub(r'\s+', ' ', n).strip()


def _parse_time_min(t: str) -> int | None:
    """Parse '10:52am' or '2:30pm (EDT)' → minutes since midnight."""
    if not t:
        return None
    t = re.sub(r'\s*\([A-Z]{2,4}\)', '', t).strip()
    m = re.match(r'(\d{1,2}):(\d{2})\s*(am|pm)', t, re.I)
    if not m:
        return None
    h, mn, ap = int(m.group(1)), int(m.group(2)), m.group(3).lower()
    if ap == 'pm' and h != 12:
        h += 12
    elif ap == 'am' and h == 12:
        h = 0
    return h * 60 + mn


def _match_task_threshold(trade: str, description: str, notes: str = "") -> dict:
    """
    §5.9 — Match trade/description/notes against 46-category threshold table.
    Fuzzy: applies regex patterns to combined lowercased text.
    Returns first match, or _FALLBACK_THRESHOLD if none.
    """
    search_text = f"{trade} {description} {notes}".lower()
    for entry in TASK_THRESHOLDS:
        if entry["pattern"] and entry["pattern"].search(search_text):
            return entry
    return _FALLBACK_THRESHOLD


def _check_threshold(actual: float, entry: dict) -> tuple[str, str] | None:
    """
    Returns ('UNDER', msg) if actual < p10,
            ('OVER',  msg) if actual >= p90,
            None if within normal range.
    Office/Admin (over_p90=0.0) skipped — handled by NB detection.
    """
    if entry is None or actual <= 0:
        return None
    if entry.get("over_p90") == 0.0:
        return None   # Office/Admin handled elsewhere

    cat   = entry["category"]
    conf  = entry["confidence"]
    c_tag = f" [{conf}]" if conf in ("LOW", "INSUFFICIENT") else ""

    if entry["under_p10"] is not None and actual < entry["under_p10"]:
        return ("UNDER",
                f"⬇ Under threshold [{cat}]: {actual:.2f}h < p10 "
                f"({entry['under_p10']:.2f}h, median {entry['median']:.2f}h){c_tag}")
    if entry["over_p90"] is not None and actual >= entry["over_p90"]:
        return ("OVER",
                f"⬆ Over threshold [{cat}]: {actual:.2f}h ≥ p90 "
                f"({entry['over_p90']:.2f}h, median {entry['median']:.2f}h){c_tag}")
    return None


def calc_billable(actual: float, non_billable: bool) -> float:
    """§5.6: NB or zero/negative hours → 0 | else min 1 hr | round UP to nearest 0.25 hr"""
    if non_billable or actual <= 0:
        return 0.0
    rounded = math.ceil(actual * 4) / 4
    return max(1.0, rounded)


def non_billable_note(notes: str) -> str:
    """Return §5.8-compliant Notes-column string for non-billable entries."""
    low = notes.lower()
    if "not billable" in low or "non-billable" in low:
        for phrase in ["not billable-", "not billable –", "not billable -",
                       "non-billable –", "non-billable -"]:
            if phrase in low:
                idx = low.index(phrase)
                return notes[idx:].split("\n")[0].strip()
        return notes.split("\n")[0].strip()
    if "office" in low or "meeting" in low or "reunión" in low or "reunion" in low:
        return "Not billable – office/meeting time"
    if "dump" in low or "vertedero" in low:
        return "Not billable – dump run"
    if "keys" in low or "llaves" in low or "entrega" in low or "drop off" in low:
        return "Not billable – key drop/pickup"
    if "clock-out" in low:
        return "Not billable – clock-out entry"
    if "these hours are not being charged" in low:
        return "Not billable – technician noted no charge"
    if "drove" in low or "drive" in low or "travel" in low:
        return "Not billable – travel time"
    return "Not billable – admin/other"


# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSER  (QB Time / TSheets payroll report)
# ─────────────────────────────────────────────────────────────────────────────

_NUMERIC_LINE_RE = re.compile(r'^[\d\s.]+$')
_FIVE_FLOATS_RE  = re.compile(
    r'^(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)$'
)
_LABELED_SUMMARY_RE = re.compile(
    r'Regular\s+(\d+\.\d+)\s+PTO\s+(\d+\.\d+)\s+OT\s+(\d+\.\d+)'
    r'\s+DT\s+(\d+\.\d+)\s+Total\s+Hours?\s+(\d+\.\d+)',
    re.IGNORECASE
)
_NON_NAME_WORDS = {
    'Regular', 'PTO', 'OT', 'DT', 'Total', 'Hours', 'Time', 'in', 'out',
    'Duration', 'Generated', 'for', 'Stonelink', 'Property', 'Management', 'Shift',
}


def _is_name_word(w: str) -> bool:
    return bool(w) and w[0].isupper() and re.match(r'^[A-Za-z]+$', w) \
           and w not in _NON_NAME_WORDS


def _extract_name_words(line: str) -> list[str]:
    words = []
    for w in line.split():
        if _is_name_word(w):
            words.append(w)
        else:
            break
    return words


def _find_name(lines: list[str], date_idx: int) -> str:
    candidate_parts: list[str] = []
    for back in range(1, 12):
        idx = date_idx - back
        if idx < 0:
            break
        line = lines[idx].strip()
        if not line:
            continue
        if line.startswith("Generated"):
            break
        if re.match(r'^\d{2}/\d{2}/\d{4}', line):
            break
        if _NUMERIC_LINE_RE.match(line):
            continue
        words = _extract_name_words(line)
        if words:
            candidate_parts = words + candidate_parts
            if len(candidate_parts) >= 2:
                break
            continue
        break
    return " ".join(candidate_parts) if candidate_parts else ""


def _rejoin_split_shifts(lines: list[str]) -> list[str]:
    """
    Rejoin shift rows split across lines by (EDT)/(EST) timezone markers.
    Handles 4-line, 3-line, and 2-line split formats.
    """
    _time_only = re.compile(r'^\d{1,2}:\d{2}[ap]m(?:\s*\([A-Z]{2,4}\))?$', re.I)
    _two_times = re.compile(
        r'^\d{1,2}:\d{2}[ap]m(?:\s*\([A-Z]{2,4}\))?\s+'
        r'\d{1,2}:\d{2}[ap]m(?:\s*\([A-Z]{2,4}\))?$', re.I)
    _two_times_float = re.compile(
        r'^\d{1,2}:\d{2}[ap]m(?:\s*\([A-Z]{2,4}\))?\s+'
        r'\d{1,2}:\d{2}[ap]m(?:\s*\([A-Z]{2,4}\))?\s+\d+\.\d+$', re.I)
    _float_only = re.compile(r'^\d+\.\d+$')

    result = []
    i = 0
    while i < len(lines):
        ln = lines[i].strip()

        # 4-line split: time / time / float / "Shift Total"
        if _time_only.match(ln):
            ahead, j = [], i + 1
            while j < len(lines) and len(ahead) < 3:
                pk = lines[j].strip(); j += 1
                if pk:
                    ahead.append(pk)
            if (len(ahead) == 3
                    and _time_only.match(ahead[0])
                    and _float_only.match(ahead[1])
                    and ahead[2] == "Shift Total"):
                result.append(f"{ln} {ahead[0]} {ahead[1]} Shift Total")
                i = j; continue

        # 3-line split: "time time" / float / "Shift Total"
        if _two_times.match(ln):
            ahead, j = [], i + 1
            while j < len(lines) and len(ahead) < 2:
                pk = lines[j].strip(); j += 1
                if pk:
                    ahead.append(pk)
            if (len(ahead) == 2
                    and _float_only.match(ahead[0])
                    and ahead[1] == "Shift Total"):
                result.append(f"{ln} {ahead[0]} Shift Total")
                i = j; continue

        # 2-line split: "time time float" / "Shift Total"
        if _two_times_float.match(ln):
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j < len(lines) and lines[j].strip() == "Shift Total":
                result.append(f"{ln} Shift Total")
                i = j + 1; continue

        result.append(ln)
        i += 1
    return result


def parse_qb_pdf(file_obj) -> list[dict]:
    """Parse TSheets Payroll PDF → list of employee dicts with shifts."""
    with pdfplumber.open(file_obj) as pdf:
        full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    return _parse_lines(full_text)


def _parse_lines(full_text: str) -> list[dict]:
    raw_lines = full_text.splitlines()
    # Strip page-footer lines BEFORE any other processing.
    # pdfplumber injects "Generated for Stonelink…" mid-employee-section when
    # a section spans a page boundary — caused premature flush and lost shifts.
    raw_lines = [l for l in raw_lines
                 if not l.strip().startswith("Generated for Stonelink")]
    lines = _rejoin_split_shifts(raw_lines)

    date_range_re = re.compile(r'(\d{2}/\d{2}/\d{4})\s+to\s+(\d{2}/\d{2}/\d{4})')
    month_re = re.compile(
        r'^(January|February|March|April|May|June|July|August|'
        r'September|October|November|December)\s+\d{1,2},\s+\d{4}'
    )
    shift_re = re.compile(
        r'^(\d{1,2}:\d{2}[ap]m(?:\s*\([A-Z]{2,4}\))?)\s+'
        r'(\d{1,2}:\d{2}[ap]m(?:\s*\([A-Z]{2,4}\))?)\s+'
        r'(\d+\.\d+)\s+Shift Total',
        re.IGNORECASE
    )
    single_float_re = re.compile(r'^\d+\.\d+$')

    employees: list[dict] = []
    emp: dict | None = None
    shift: dict | None = None
    cur_date = cur_date_iso = ""
    reading_notes = False
    notes_buf = ""
    expect: dict = {}

    def _flush_shift():
        nonlocal shift, reading_notes, notes_buf
        if shift and emp is not None:
            if reading_notes:
                shift["notes"] = notes_buf.strip()
                shift["meld_ids"] = extract_meld_ids(shift["notes"])
                shift["non_billable"] = is_non_billable(shift["notes"])
            emp["shifts"].append(shift)
        shift = None
        reading_notes = False
        notes_buf = ""

    def _flush_emp():
        _flush_shift()
        if emp and emp.get("name"):
            employees.append(emp)

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        m = date_range_re.match(line)
        if m:
            name = _find_name(lines, i)
            summary_from_date_line = _FIVE_FLOATS_RE.search(
                line[len(m.group(0)):].strip()
            )
            prev_five = None
            if not summary_from_date_line and i > 0:
                prev_five = _FIVE_FLOATS_RE.match(lines[i - 1].strip())

            _flush_emp()
            emp = dict(name=name, period_start=m.group(1), period_end=m.group(2),
                       regular=0.0, pto=0.0, ot=0.0, dt=0.0, total_hours=0.0,
                       days={}, shifts=[])
            expect = {"regular": True}

            def _apply_five(mx):
                emp["regular"]     = float(mx.group(1))
                emp["pto"]         = float(mx.group(2))
                emp["ot"]          = float(mx.group(3))
                emp["dt"]          = float(mx.group(4))
                emp["total_hours"] = float(mx.group(5))
                expect.clear()

            if summary_from_date_line:
                _apply_five(summary_from_date_line)
            elif prev_five:
                _apply_five(prev_five)
            else:
                if i + 1 < len(lines):
                    nxt = lines[i + 1].strip()
                    m5 = _FIVE_FLOATS_RE.match(nxt)
                    if m5:
                        _apply_five(m5); i += 2; continue
                    ml = _LABELED_SUMMARY_RE.search(nxt)
                    if ml:
                        emp["regular"]     = float(ml.group(1))
                        emp["pto"]         = float(ml.group(2))
                        emp["ot"]          = float(ml.group(3))
                        emp["dt"]          = float(ml.group(4))
                        emp["total_hours"] = float(ml.group(5))
                        expect.clear()
                        i += 2; continue

            i += 1; continue

        if emp is None:
            i += 1; continue

        for lbl, key in [("Regular","regular"), ("PTO","pto"),
                          ("OT","ot"), ("DT","dt"), ("Total Hours","total_hours")]:
            if line == lbl:
                expect[key] = True
                break
        else:
            for key in list(expect):
                if expect.get(key) and single_float_re.match(line):
                    emp[key] = float(line)
                    expect[key] = False
                    break

        if not emp["total_hours"] and _FIVE_FLOATS_RE.match(line):
            m5 = _FIVE_FLOATS_RE.match(line)
            emp["regular"]     = float(m5.group(1))
            emp["pto"]         = float(m5.group(2))
            emp["ot"]          = float(m5.group(3))
            emp["dt"]          = float(m5.group(4))
            emp["total_hours"] = float(m5.group(5))
            expect = {}
            i += 1; continue

        if not emp["total_hours"]:
            ml = _LABELED_SUMMARY_RE.search(line)
            if ml:
                emp["regular"]     = float(ml.group(1))
                emp["pto"]         = float(ml.group(2))
                emp["ot"]          = float(ml.group(3))
                emp["dt"]          = float(ml.group(4))
                emp["total_hours"] = float(ml.group(5))
                expect = {}
                i += 1; continue

        if month_re.match(line):
            _flush_shift()
            parts = line.split()
            cur_date = " ".join(parts[:3]) if len(parts) >= 3 else line
            cur_date_iso = _date_iso(cur_date)
            if len(parts) == 4 and single_float_re.match(parts[3]):
                emp["days"][cur_date] = float(parts[3])
                i += 1; continue
            if i + 1 < len(lines) and single_float_re.match(lines[i+1].strip()):
                emp["days"][cur_date] = float(lines[i+1].strip())
                i += 2; continue
            i += 1; continue

        ms = shift_re.match(line)
        if ms:
            _flush_shift()
            shift = dict(
                date=cur_date, date_iso=cur_date_iso,
                time_in=ms.group(1).replace("(EDT)","").strip(),
                time_out=ms.group(2).replace("(EDT)","").strip(),
                duration=float(ms.group(3)),
                notes="", meld_ids=[], non_billable=False,
            )
            i += 1; continue

        if line.startswith("NOTES:") and shift:
            reading_notes = True
            notes_buf = line[6:].strip()
            i += 1; continue

        if reading_notes and shift:
            is_structural = (month_re.match(line)
                             or shift_re.match(line)
                             or line.startswith("Generated")
                             or line.startswith("NOTES:"))
            if not is_structural:
                if line:
                    notes_buf += " " + line
                i += 1; continue
            else:
                shift["notes"] = notes_buf.strip()
                shift["meld_ids"] = extract_meld_ids(shift["notes"])
                shift["non_billable"] = is_non_billable(shift["notes"])
                reading_notes = False
                notes_buf = ""

        if line.startswith("Generated for"):
            _flush_emp()
            emp = None
            i += 1; continue

        i += 1

    _flush_emp()
    return employees


def _date_iso(s: str) -> str:
    MONTHS = {"January":"01","February":"02","March":"03","April":"04",
               "May":"05","June":"06","July":"07","August":"08",
               "September":"09","October":"10","November":"11","December":"12"}
    m = re.match(r'(\w+)\s+(\d+),\s+(\d+)', s)
    if not m:
        return ""
    return f"{m.group(3)}-{MONTHS.get(m.group(1),'00')}-{m.group(2).zfill(2)}"


# ─────────────────────────────────────────────────────────────────────────────
# CSV / MELD LOADER
# ─────────────────────────────────────────────────────────────────────────────

def load_melds(file_obj) -> tuple[pd.DataFrame, list[str]]:
    """
    §4.1 / §4.3 (SOP v2.0): Property Meld Work Log Summary is PRIMARY source.
    Validates all 7 required columns before proceeding.
    Returns (DataFrame, list_of_warning_strings).
    """
    df = pd.read_csv(file_obj, dtype=str)
    df.columns = df.columns.str.strip()

    if "Meld Number" in df.columns:
        df = df.rename(columns={"Meld Number": "Meld"})
    if "Address line 1" in df.columns and "Address Line 1" not in df.columns:
        df = df.rename(columns={"Address line 1": "Address Line 1"})

    if "Meld" not in df.columns:
        raise ValueError("CSV must contain a 'Meld' or 'Meld Number' column.")

    # §4.3 — Validate 7 required columns
    warnings_out: list[str] = []
    effective_cols = set(df.columns)
    # "Check In" is accepted as equivalent to "Check-In Hours" for validation
    if "Check In" in effective_cols:
        effective_cols.add("Check-In Hours")
    missing = REQUIRED_CSV_COLS - effective_cols
    if missing:
        warnings_out.append(
            f"⚠ **§4.3 CONTROL: Missing required CSV columns: "
            f"{', '.join(sorted(missing))}** — "
            f"confirm all 7 approved columns were exported from Property Meld."
        )

    df["Meld"] = df["Meld"].str.strip().str.upper()
    for col in ["Hours", "Check-In Hours", "Total Labor Hours"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    return df.reset_index(drop=True), warnings_out


def _clean(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none") else s


def meld_lookup(row: pd.Series | None, field: str, default="") -> str:
    if row is None:
        return default
    val = row.get(field, default)
    if isinstance(val, pd.Series):
        val = val.iloc[0] if len(val) > 0 else default
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return str(default).strip()
    return str(val).strip()


# ─────────────────────────────────────────────────────────────────────────────
# BUILD ROW DATA
# ─────────────────────────────────────────────────────────────────────────────

def build_recap_rows(employees: list[dict], melds_df: pd.DataFrame) -> list[dict]:
    """
    SOP v2.0 §4.1: Property Meld CSV is the FOUNDATION.
    One output row per CSV work log entry + unmatched TSheets shifts as 'No Meld'.
    All 8 §5.9 flag checks applied. Per-category threshold checks applied.
    """

    # ── Build TSheets lookups ─────────────────────────────────────────────
    ts_name_meld_date: dict[tuple, dict] = {}
    ts_name_meld:      dict[tuple, list] = {}
    ts_meld_only:      dict[str,   list] = {}
    ts_name_date:      dict[tuple, list] = {}
    all_shifts: list[tuple[dict, str]] = []   # (shift_dict, emp_name)

    for emp in employees:
        name_n = _norm_name(emp["name"])
        for sh in emp["shifts"]:
            all_shifts.append((sh, emp["name"]))
            ts_name_date.setdefault((name_n, sh["date_iso"]), []).append(sh)
            for mid in sh["meld_ids"]:
                mid_u = mid.upper()
                ts_name_meld_date[(name_n, mid_u, sh["date_iso"])] = sh
                ts_name_meld.setdefault((name_n, mid_u), []).append(sh)
                ts_meld_only.setdefault(mid_u, []).append(sh)

    ts_matched_ids: set[int] = set()
    rows: list[dict] = []

    # ── Process each CSV row ──────────────────────────────────────────────
    for _, csv_row in melds_df.iterrows():
        meld    = _clean(csv_row.get("Meld", "")).upper()
        agent   = _clean(csv_row.get("Agent", ""))
        agent_n = _norm_name(agent)

        checkin_raw = _clean(
            csv_row.get("Check In") or csv_row.get("Check-In Hours") or ""
        )
        date_iso, date_display, time_in_display = "", "", ""
        if checkin_raw:
            try:
                dt = pd.to_datetime(checkin_raw)
                date_iso        = dt.strftime("%Y-%m-%d")
                date_display    = f"{dt.strftime('%B')} {dt.day}, {dt.year}"
                hour            = dt.strftime("%I").lstrip("0") or "12"
                time_in_display = f"{hour}:{dt.strftime('%M%p').lower()}"
            except Exception:
                date_display = checkin_raw

        # Tier 1: exact normalized name + meld + date
        tshift = ts_name_meld_date.get((agent_n, meld, date_iso))

        # Tier 2: normalized name + meld (any date)
        if not tshift:
            cands = ts_name_meld.get((agent_n, meld), [])
            tshift = next((c for c in cands if c["date_iso"] == date_iso), None)
            if not tshift and cands:
                tshift = cands[0]

        # Tier 3: Meld ID only (handles name spelling differences)
        if not tshift and meld:
            meld_cands = ts_meld_only.get(meld, [])
            tshift = next((c for c in meld_cands if c["date_iso"] == date_iso), None)
            if not tshift and meld_cands:
                tshift = meld_cands[0]

        # Tier 3.5: same tech + same date + closest duration (within 0.25 hrs)
        if not tshift and date_iso and agent_n:
            csv_hrs_raw = csv_row.get("Check-In Hours") or csv_row.get("Hours") or 0
            try:
                csv_hrs_f = float(csv_hrs_raw)
            except (ValueError, TypeError):
                csv_hrs_f = 0.0
            if csv_hrs_f > 0:
                day_shifts = ts_name_date.get((agent_n, date_iso), [])
                best, best_diff = None, 999.0
                for s in day_shifts:
                    diff = abs(s["duration"] - csv_hrs_f)
                    if diff < best_diff:
                        best_diff = diff
                        best = s
                if best and best_diff <= 0.25:
                    tshift = best

        if tshift:
            ts_matched_ids.add(id(tshift))

        # ── Hours resolution (§4.1) ───────────────────────────────────────
        # Property Meld CSV Hours is the PRIMARY per-work-order time and is
        # authoritative.  The matched TSheets shift is used ONLY to verify the
        # tech clocked in and to source the technician's notes — NOT for hours.
        # (Using TSheets duration here double-counted shifts that span multiple
        #  CSV rows and inflated zero-hour admin rows.)
        csv_hrs = (csv_row.get("Hours")
                   if csv_row.get("Hours") not in (None, "")
                   else (csv_row.get("Check-In Hours")
                         or csv_row.get("Total Labor Hours") or 0))
        try:
            actual = float(csv_hrs)
        except (ValueError, TypeError):
            actual = 0.0

        ts_duration = tshift["duration"] if tshift else None
        if tshift:
            time_in   = tshift["time_in"] or time_in_display
            raw_notes = tshift["notes"]
            nb        = tshift["non_billable"]
            flat_b    = is_flat_bill(tshift["notes"])
        else:
            time_in   = time_in_display
            raw_notes = ""
            nb        = False
            flat_b    = False

        # §5.6 — RC and Protected trade checks
        rc        = is_rc(raw_notes)
        trade     = (_clean(csv_row.get("Title"))
                     or _clean(csv_row.get("Work Category"))
                     or _clean(csv_row.get("Work Type")))
        desc      = (_clean(csv_row.get("Description"))
                     or _clean(csv_row.get("Meld Description")))
        protected = is_protected_trade(trade, desc)

        # §5.6 — Block NB override for protected trades
        nb_overridden = False
        if nb and protected:
            nb = False
            nb_overridden = True

        paying   = actual
        billable = calc_billable(actual, nb)

        address = (_clean(csv_row.get("Address Line 1"))
                   or _clean(csv_row.get("Address line 1"))
                   or _clean(csv_row.get("Property Name")))
        unit   = _clean(csv_row.get("Unit"))
        status = _clean(csv_row.get("Meld Status"))

        # §5.9 — Flag checks
        is_fifteen = _is_15_circle(address)
        under_01   = (0 < actual < 0.10)

        # Turnover detection — skip threshold check (handled separately)
        is_turn = is_turnover(trade, desc, raw_notes)

        # Threshold check (billable entries only, non-turnover only)
        thresh_result = None
        if actual > 0 and not nb and not is_turn:
            t_entry = _match_task_threshold(trade, desc, raw_notes)
            thresh_result = _check_threshold(actual, t_entry)

        # §5.8 — Notes (approved conventions only)
        flag_parts: list[str] = []
        if flat_b:
            flag_parts.append("Not billable- to be flat bill")
        elif rc:
            flag_parts.append("RC – resident charge")
        elif nb:
            flag_parts.append(non_billable_note(raw_notes))
        if nb_overridden:
            flag_parts.append(
                "REVIEW NEEDED – §5.6: Protected trade NB blocked; manager auth required"
            )
        if not tshift and not actual:
            flag_parts.append("REVIEW NEEDED – no hours in TSheets or CSV")
        if thresh_result:
            flag_parts.append(thresh_result[1])

        j_note = " | ".join(filter(None, flag_parts + [desc]))
        has_ps = bool(_PS_NOTE_RE.search(j_note))

        rows.append({
            "MWO":            meld or "No Meld",
            "Tech":           agent,
            "Address":        address,
            "Unit":           unit,
            "Check-In":       time_in,
            "Actual":         actual,
            "Paying":         paying,
            "Billable":       billable,
            "Notes":          j_note,
            "Trade":          trade,
            "Work Performed": desc,
            "_employee":      agent,
            "_date":          date_display,
            "_date_iso":      date_iso,
            "_time_out":      tshift["time_out"] if tshift else "",
            "_meld_ids":      [meld] if meld else [],
            "_non_bill":      nb,
            "_flat_bill":     flat_b,
            "_rc":            rc,
            "_protected":     protected,
            "_nb_overridden": nb_overridden,
            "_is_15_circle":  is_fifteen,
            "_under_01":      under_01,
            "_thresh_result": thresh_result,
            "_has_ps_note":   has_ps,
            "_nb_missing_note": nb and not flag_parts,
            "_meld_found":    True,
            "_tsheets_matched": tshift is not None,
            "_no_mwo":        not meld,
            "_no_trade":      not trade,
            "_no_raw_notes":  not raw_notes.strip(),
            "_raw_notes":     raw_notes,
            "_status":        status,
            "_no_meld_entry": False,
            "_possible_dup":  False,
            "_overlap_flag":  False,
            "_turnover":      is_turn,
            "_non_labor":     actual <= 0,   # admin/contact/scoping (no field time)
            "_ts_duration":   ts_duration,   # matched TSheets shift hrs (verify only)
            "_approval":      "Approved",   # set in post-process below
        })

    # ── §7.3 — Unmatched TSheets shifts → 'No Meld' rows ─────────────────
    for sh, emp_name in all_shifts:
        if id(sh) in ts_matched_ids or sh["duration"] <= 0:
            continue
        nb_sh = sh["non_billable"]
        rc_sh = is_rc(sh["notes"])
        rows.append({
            "MWO":            "No Meld",
            "Tech":           emp_name,
            "Address":        "",
            "Unit":           "",
            "Check-In":       sh["time_in"],
            "Actual":         sh["duration"],
            "Paying":         sh["duration"],
            "Billable":       0.0 if nb_sh else calc_billable(sh["duration"], nb_sh),
            "Notes":          (f"⚠ No Property Meld match – {sh['notes'][:80]}"
                               if sh["notes"] else "⚠ No Property Meld match"),
            "Trade":          "",
            "Work Performed": sh["notes"],
            "_employee":      emp_name,
            "_date":          sh["date"],
            "_date_iso":      sh["date_iso"],
            "_time_out":      sh["time_out"],
            "_meld_ids":      sh["meld_ids"],
            "_non_bill":      nb_sh,
            "_flat_bill":     False,
            "_rc":            rc_sh,
            "_protected":     False,
            "_nb_overridden": False,
            "_is_15_circle":  False,
            "_under_01":      0 < sh["duration"] < 0.10,
            "_thresh_result": None,
            "_has_ps_note":   False,
            "_nb_missing_note": False,
            "_meld_found":    True,
            "_tsheets_matched": True,
            "_no_mwo":        True,
            "_no_trade":      True,
            "_no_raw_notes":  not sh["notes"].strip(),
            "_raw_notes":     sh["notes"],
            "_status":        "",
            "_no_meld_entry": True,
            "_possible_dup":  False,
            "_overlap_flag":  False,
            "_turnover":      is_turnover("", sh["notes"]),
            "_non_labor":     False,
            "_ts_duration":   sh["duration"],
            "_approval":      "Approved",   # set in post-process below
        })

    # ── §5.9 Post-process: Possible Duplicate ─────────────────────────────
    seen_keys: dict[tuple, list[int]] = {}
    for i, r in enumerate(rows):
        if r["MWO"] and r["MWO"] != "No Meld" and r["_date_iso"]:
            seen_keys.setdefault((r["MWO"], r["Tech"], r["_date_iso"]), []).append(i)
    for indices in seen_keys.values():
        if len(indices) > 1:
            for idx in indices:
                rows[idx]["_possible_dup"] = True

    # ── §5.9 Post-process: Overlapping Tech Conflict ──────────────────────
    tech_day: dict[tuple, list[tuple]] = {}
    for i, r in enumerate(rows):
        if r["_tsheets_matched"] and r["_date_iso"] and r["Actual"] > 0:
            key = (r["Tech"].lower(), r["_date_iso"])
            s_min = _parse_time_min(r["Check-In"])
            if s_min is not None:
                e_min = s_min + int(r["Actual"] * 60)
                tech_day.setdefault(key, []).append((s_min, e_min, i))

    for entries in tech_day.values():
        if len(entries) < 2:
            continue
        for a in range(len(entries)):
            for b in range(a + 1, len(entries)):
                s1, e1, i1 = entries[a]
                s2, e2, i2 = entries[b]
                if s1 < e2 and s2 < e1:   # intervals overlap
                    rows[i1]["_overlap_flag"] = True
                    rows[i2]["_overlap_flag"] = True

    # ── Approval Status: any tripped flag → "Review Required" ─────────────
    # Default is "Approved"; ANY of the review-worthy flags overrides it.
    # Flat-bill is a known pre-approved convention and does NOT trip review.
    for r in rows:
        if _row_needs_review(r):
            r["_approval"] = "Review Required"

    return rows


def _row_needs_review(r: dict) -> bool:
    """Return True if any flag should override default 'Approved' status."""
    # Non-labor admin/contact rows (0 hrs) carry no billable risk → stay Approved.
    if r.get("_non_labor"):
        return False
    return bool(
        r.get("_thresh_result") is not None          # UNDER or OVER
        or not r.get("_tsheets_matched", True)        # No TSheets match
        or r.get("_no_meld_entry")                    # Unmatched shift
        or r.get("_nb_missing_note")                  # NB without note (§5.5)
        or r.get("_nb_overridden")                    # Protected trade NB blocked
        or r.get("_is_15_circle")                     # 15 Circle St (separate billing)
        or r.get("_rc")                               # Resident Charge
        or r.get("_has_ps_note")                      # PS: instruction note
        or r.get("_possible_dup")                     # Possible duplicate
        or r.get("_overlap_flag")                     # Overlapping shift
        or r.get("_under_01")                         # < 0.10 hrs
        or (r["Paying"] > r["Billable"] and r["Billable"] > 0)  # Paying > Billable
    )


def build_payroll_rows(employees: list[dict], recap_rows: list[dict] | None = None) -> list[dict]:
    """
    §5.11 SOP v2.0 — Payroll Hours Spreadsheet
    PAYROLL pays the employee for hours CLOCKED in TSheets (QB Time payroll
    summary) — NOT the client-billable hours from the recap.  Regular/OT come
    straight from the QB Time summary (which already splits at 40h); Vacation =
    TSheets PTO.  Holiday + Sick/Flex are filled in manually.
    (recap_rows kept for signature compatibility; no longer used for hours.)
    """
    rows = []
    for emp in employees:
        full  = emp["name"].strip()
        parts = full.split()
        last  = parts[-1] if parts else full
        first = " ".join(parts[:-1]) if len(parts) > 1 else ""

        regular   = emp.get("regular", 0.0)
        ot        = emp.get("ot", 0.0)
        # QB Time sometimes reports all worked hours under Regular without
        # splitting OT — apply the 40h split defensively if OT wasn't provided.
        if ot == 0.0 and regular > OT_THRESHOLD:
            ot      = regular - OT_THRESHOLD
            regular = OT_THRESHOLD
        vacation  = emp.get("pto", 0.0)
        holiday   = 0.0
        sick_flex = 0.0
        total     = regular + vacation + holiday + sick_flex + ot

        rows.append({
            "Last Name":  last,
            "First Name": first,
            "Regular":    regular,
            "Vacation":   vacation,
            "Holiday":    holiday,
            "Sick/Flex":  sick_flex,
            "Overtime":   ot,
            "Total":      total,
            "_ot_flag":   ot > 0,
            "_pto_flag":  vacation > 0,
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _c(ws, r, c, val="", fill=None, font=None, align="left", fmt=None,
       bold=False, wrap=False, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    if fill:   cell.fill = fill
    if font:   cell.font = font
    elif bold: cell.font = BOLD_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt:    cell.number_format = fmt
    if border: cell.border = BOX
    return cell


def build_excel(recap_rows: list[dict], payroll_rows: list[dict],
                employees: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_weekly_recap(wb, recap_rows, employees)
    _sheet_payroll_hours(wb, payroll_rows)
    _sheet_flag_summary(wb, recap_rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tab 1: Weekly Recap ──────────────────────────────────────────────────────

RECAP_COLS = [
    ("A", "MWO",            16),
    ("B", "Tech",           20),
    ("C", "Address",        28),
    ("D", "Unit",           10),
    ("E", "Check-In",       10),
    ("F", "Actual",         10),
    ("G", "Paying",         10),
    ("H", "Billable",       10),
    ("I", "Status",         16),
    ("J", "Notes",          32),
    ("K", "Trade",          18),
    ("L", "Work Performed", 38),
]

STATUS_REVIEW_FILL  = PatternFill("solid", fgColor="FCE4EC")   # subtle pink — review needed
STATUS_APPROVED_FILL = PatternFill("solid", fgColor="E8F5E9")  # subtle green — approved


def _sheet_weekly_recap(wb: Workbook, rows: list[dict], employees: list[dict]):
    ws = wb.create_sheet("Weekly Recap")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:L1")
    t = ws.cell(row=1, column=1, value="Weekly Recap – Planet Synergy PM Report")
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = DARK_BLUE
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    period = ""
    if employees:
        period = f"Pay Period: {employees[0]['period_start']}  –  {employees[0]['period_end']}"
    ws.merge_cells("A2:L2")
    pc = ws.cell(row=2, column=1, value=period)
    pc.font = Font(bold=True, size=10, color="FFFFFF")
    pc.fill = MED_BLUE
    pc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    for idx, (_, hdr, width) in enumerate(RECAP_COLS, 1):
        ltr = get_column_letter(idx)
        ws.column_dimensions[ltr].width = width
        _c(ws, 3, idx, hdr, fill=DARK_BLUE, font=WHITE_FONT, align="center", wrap=True)
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"

    # Categorize into four mutually-exclusive sections so each total is clean:
    #   1. Standard billing  – matched Property Meld work, not 15 Circle St
    #   2. 15 Circle St       – billed independently (§5.9/§6)
    #   3. No-Meld            – unmatched TSheets time, no work order (§7.3)
    #   4. Non-Labor / Admin  – 0-hour CSV entries (contact/scoping/status)
    nonlabor_rows   = [r for r in rows if r.get("_non_labor")]
    nomeld_rows     = [r for r in rows if r.get("_no_meld_entry")
                       and not r.get("_non_labor")]
    labor_rows      = [r for r in rows if not r.get("_non_labor")
                       and not r.get("_no_meld_entry")]
    standard_rows   = [r for r in labor_rows if not r.get("_is_15_circle")]
    circle_rows     = [r for r in labor_rows if r.get("_is_15_circle")]
    sorted_standard = sorted(standard_rows,
                             key=lambda r: (r["Tech"], r["_date_iso"], r["Check-In"]))
    sorted_circle   = sorted(circle_rows,
                             key=lambda r: (r["Tech"], r["_date_iso"], r["Check-In"]))
    sorted_nomeld   = sorted(nomeld_rows,
                             key=lambda r: (r["Tech"], r["_date_iso"], r["Check-In"]))
    sorted_nonlabor = sorted(nonlabor_rows,
                             key=lambda r: (r["Tech"], r["_date_iso"], r["Check-In"]))

    flags: list[dict] = []

    def _write_employee_group(emp_name: str, emp_rows: list[dict],
                              data_row: int) -> int:
        fill_idx = 0
        f_start  = data_row

        for r in emp_rows:
            base_fill = ALT_FILL if fill_idx % 2 == 0 else WHITE_FILL
            fill_idx += 1

            ts_missing  = not r.get("_tsheets_matched", True) or r.get("_no_meld_entry")
            nb_no_note  = r.get("_nb_missing_note", False)
            is_fifteen  = r.get("_is_15_circle", False)
            rc_entry    = r.get("_rc", False)
            has_ps      = r.get("_has_ps_note", False)
            pay_gt_bill = r["Paying"] > r["Billable"] and r["Billable"] > 0
            t_res       = r.get("_thresh_result")
            flat_bill   = r.get("_flat_bill", False)

            # Priority: RED > ORANGE > LAVENDER > TEAL > YELLOW > AMBER > OVER > UNDER > PURPLE
            if ts_missing:
                row_fill = RED_FILL
                flags.append({"type": "🔴 No TSheets Match / No Meld",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "hours": r["Actual"]})
            elif nb_no_note:
                row_fill = ORANGE_FILL
                flags.append({"type": "🟠 Non-Billable – No Note (§5.5 Error)",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "hours": r["Actual"]})
            elif is_fifteen:
                row_fill = LAVENDER_FILL
            elif rc_entry:
                row_fill = TEAL_FILL
                flags.append({"type": "🔵 RC – Resident Charge",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "hours": r["Actual"]})
            elif has_ps:
                row_fill = YELLOW_FILL
                flags.append({"type": "🟡 PS: Note – Planet Synergy Action",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "hours": r["Actual"]})
            elif pay_gt_bill:
                row_fill = AMBER_FILL
                flags.append({"type": "⚠ Paying > Billable – Manager Review",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"],
                              "actual": r["Actual"], "billable": r["Billable"]})
            elif t_res and t_res[0] == "OVER":
                row_fill = OVER_FILL
                flags.append({"type": "⬆ Threshold OVER",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "detail": t_res[1]})
            elif t_res and t_res[0] == "UNDER":
                row_fill = UNDER_FILL
                flags.append({"type": "⬇ Threshold UNDER",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "detail": t_res[1]})
            elif flat_bill:
                row_fill = PURPLE_FILL
            else:
                row_fill = base_fill

            if r.get("_possible_dup"):
                flags.append({"type": "🔁 Possible Duplicate",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "hours": r["Actual"]})
            if r.get("_overlap_flag"):
                flags.append({"type": "⏱ Overlapping Shift Conflict",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "hours": r["Actual"]})
            if r.get("_under_01"):
                flags.append({"type": "⚠ Under 0.10 Hours",
                              "employee": emp_name, "meld": r["MWO"],
                              "date": r["_date"], "hours": r["Actual"]})

            approval = r.get("_approval", "Approved")
            status_text = "⚠ Review" if approval == "Review Required" else "✓ Approved"

            vals = [r["MWO"], r["Tech"], r["Address"], r["Unit"], r["Check-In"],
                    r["Actual"], r["Paying"], r["Billable"], status_text,
                    r["Notes"], r["Trade"], r["Work Performed"]]
            for c_idx, val in enumerate(vals, 1):
                is_num = c_idx in (6, 7, 8)
                # Status column (9) keeps its own fill so it stays visible
                # regardless of the row's flag-color fill
                if c_idx == 9:
                    cell_fill = (STATUS_REVIEW_FILL
                                 if approval == "Review Required"
                                 else STATUS_APPROVED_FILL)
                    _c(ws, data_row, c_idx, val, fill=cell_fill,
                       align="center", bold=True)
                else:
                    _c(ws, data_row, c_idx, val, fill=row_fill,
                       align="center" if is_num else "left",
                       fmt="0.00" if is_num else None,
                       wrap=(c_idx in (3, 10, 12)))
            data_row += 1

        # Subtotal row
        f_end = data_row - 1
        _c(ws, data_row, 1, f"Subtotal – {emp_name}", fill=GREEN_FILL,
           bold=True, align="left")
        ws.merge_cells(f"A{data_row}:E{data_row}")
        ws.cell(row=data_row, column=1).fill = GREEN_FILL
        for ci, ltr in [(6,"F"), (7,"G"), (8,"H")]:
            cell = ws.cell(row=data_row, column=ci,
                           value=f"=SUM({ltr}{f_start}:{ltr}{f_end})")
            cell.fill = GREEN_FILL; cell.font = BOLD_FONT
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BOX
        for ci in [9, 10, 11, 12]:
            ws.cell(row=data_row, column=ci).fill = GREEN_FILL
            ws.cell(row=data_row, column=ci).border = BOX
        ws.row_dimensions[data_row].height = 16
        data_row += 1
        _c(ws, data_row, 1, "", fill=WHITE_FILL, border=False)
        return data_row + 1

    # ── Standard billing section ──────────────────────────────────────────
    data_row = 4
    emp_groups: dict[str, list[dict]] = defaultdict(list)
    for r in sorted_standard:
        emp_groups[r["Tech"]].append(r)
    for emp_name, emp_rows in emp_groups.items():
        data_row = _write_employee_group(emp_name, emp_rows, data_row)

    # Grand total — standard
    gt_row = data_row
    ws.merge_cells(f"A{gt_row}:E{gt_row}")
    _c(ws, gt_row, 1, "GRAND TOTAL – STANDARD BILLING",
       fill=DARK_BLUE, font=WHITE_FONT, align="left")
    for ci in range(2, 6):
        ws.cell(row=gt_row, column=ci).fill = DARK_BLUE
        ws.cell(row=gt_row, column=ci).border = BOX
    for ci, ltr in [(6,"F"), (7,"G"), (8,"H")]:
        # Subtotal label lives in column A — filter on A (not B) so subtotal
        # rows are excluded and not double-counted into the grand total.
        cell = ws.cell(row=gt_row, column=ci,
                       value=f"=SUMIF(A4:A{gt_row-1},\"<>Subtotal*\","
                             f"{ltr}4:{ltr}{gt_row-1})")
        cell.fill = DARK_BLUE; cell.font = WHITE_FONT
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX
    for ci in [9, 10, 11, 12]:
        ws.cell(row=gt_row, column=ci).fill = DARK_BLUE
        ws.cell(row=gt_row, column=ci).border = BOX
    ws.row_dimensions[gt_row].height = 18
    data_row = gt_row + 2

    # ── 15 Circle St section (§5.9 — separate billing) ───────────────────
    if circle_rows:
        # Section header
        ws.merge_cells(f"A{data_row}:L{data_row}")
        hdr = ws.cell(row=data_row, column=1,
                      value="⚠ 15 CIRCLE ST (RUMFORD, RI) — BILLED INDEPENDENTLY (§5.9/§6)")
        hdr.font = Font(bold=True, size=11, color="FFFFFF")
        hdr.fill = PatternFill("solid", fgColor="6A1B9A")
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[data_row].height = 18
        data_row += 1

        cs_data_start = data_row
        circle_groups: dict[str, list[dict]] = defaultdict(list)
        for r in sorted_circle:
            circle_groups[r["Tech"]].append(r)
        for emp_name, emp_rows in circle_groups.items():
            data_row = _write_employee_group(emp_name, emp_rows, data_row)

        # 15 Circle St total
        cs_data_end = data_row - 1
        ct_row = data_row
        ws.merge_cells(f"A{ct_row}:E{ct_row}")
        _c(ws, ct_row, 1, "15 CIRCLE ST TOTAL (Separate Billing)",
           fill=DARK_BLUE, font=WHITE_FONT)
        ws.cell(row=ct_row, column=1).fill = DARK_BLUE
        for ci in range(2, 6):
            ws.cell(row=ct_row, column=ci).fill = DARK_BLUE
            ws.cell(row=ct_row, column=ci).border = BOX
        for ci, ltr in [(6,"F"), (7,"G"), (8,"H")]:
            cell = ws.cell(row=ct_row, column=ci,
                           value=f"=SUMIF(A{cs_data_start}:A{cs_data_end},"
                                 f"\"<>Subtotal*\","
                                 f"{ltr}{cs_data_start}:{ltr}{cs_data_end})")
            cell.fill = DARK_BLUE; cell.font = WHITE_FONT
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BOX
        for ci in [9, 10, 11, 12]:
            ws.cell(row=ct_row, column=ci).fill = DARK_BLUE
            ws.cell(row=ct_row, column=ci).border = BOX
        ws.row_dimensions[ct_row].height = 18
        data_row = ct_row + 2

    # ── No-Meld section (unmatched TSheets time — §7.3) ───────────────────
    if sorted_nomeld:
        ws.merge_cells(f"A{data_row}:L{data_row}")
        hdr = ws.cell(row=data_row, column=1,
                      value="🔴 NO-MELD — TSheets time with NO Property Meld work order "
                            "(escalate to manager; NOT part of standard billing) (§7.3)")
        hdr.font = Font(bold=True, size=11, color="FFFFFF")
        hdr.fill = PatternFill("solid", fgColor="B71C1C")
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[data_row].height = 18
        data_row += 1

        nm_data_start = data_row
        nm_groups: dict[str, list[dict]] = defaultdict(list)
        for r in sorted_nomeld:
            nm_groups[r["Tech"]].append(r)
        for emp_name, emp_rows in nm_groups.items():
            data_row = _write_employee_group(emp_name, emp_rows, data_row)

        nm_data_end = data_row - 1
        nt_row = data_row
        ws.merge_cells(f"A{nt_row}:E{nt_row}")
        _c(ws, nt_row, 1, "NO-MELD TOTAL (Unmatched — escalate, not billed)",
           fill=DARK_BLUE, font=WHITE_FONT)
        ws.cell(row=nt_row, column=1).fill = DARK_BLUE
        for ci in range(2, 6):
            ws.cell(row=nt_row, column=ci).fill = DARK_BLUE
            ws.cell(row=nt_row, column=ci).border = BOX
        for ci, ltr in [(6,"F"), (7,"G"), (8,"H")]:
            cell = ws.cell(row=nt_row, column=ci,
                           value=f"=SUMIF(A{nm_data_start}:A{nm_data_end},"
                                 f"\"<>Subtotal*\","
                                 f"{ltr}{nm_data_start}:{ltr}{nm_data_end})")
            cell.fill = DARK_BLUE; cell.font = WHITE_FONT
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BOX
        for ci in [9, 10, 11, 12]:
            ws.cell(row=nt_row, column=ci).fill = DARK_BLUE
            ws.cell(row=nt_row, column=ci).border = BOX
        ws.row_dimensions[nt_row].height = 18
        data_row = nt_row + 2

    # ── Non-Labor / Admin section (0-hour CSV entries) ────────────────────
    if sorted_nonlabor:
        ws.merge_cells(f"A{data_row}:L{data_row}")
        hdr = ws.cell(row=data_row, column=1,
                      value="NON-LABOR / ADMIN ENTRIES — resident contact, scoping, "
                            "status updates (0 field hours; not billed)")
        hdr.font = Font(bold=True, size=11, color="FFFFFF")
        hdr.fill = PatternFill("solid", fgColor="616161")
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[data_row].height = 18
        data_row += 1

        nl_groups: dict[str, list[dict]] = defaultdict(list)
        for r in sorted_nonlabor:
            nl_groups[r["Tech"]].append(r)
        for emp_name, emp_rows in nl_groups.items():
            data_row = _write_employee_group(emp_name, emp_rows, data_row)
        data_row += 1

    # Legend
    leg_row = data_row
    ws.cell(row=leg_row, column=1, value="Legend:").font = BOLD_FONT
    legend_items = [
        (STATUS_APPROVED_FILL, "Status: ✓ Approved – default; no flags tripped"),
        (STATUS_REVIEW_FILL,   "Status: ⚠ Review – any flag overrides default approval"),
        (RED_FILL,      "No TSheets match / No Meld – verify work order (§4.1/§7.3)"),
        (ORANGE_FILL,   "Non-billable missing note – ACTIONABLE ERROR (§5.5)"),
        (LAVENDER_FILL, "15 Circle St – billed independently (§5.9/§6)"),
        (TEAL_FILL,     "RC – Resident Charge (§5.6)"),
        (YELLOW_FILL,   "PS: note – Planet Synergy action required (§5.8)"),
        (AMBER_FILL,    "Paying > Billable – manager review required (§6)"),
        (OVER_FILL,     "Over p90 threshold – review with manager (§5.9)"),
        (UNDER_FILL,    "Under p10 threshold – possible missing time (§5.9)"),
        (PURPLE_FILL,   "Flat-bill entry – pre-approved; does NOT trip review"),
        (GREEN_FILL,    "Technician subtotal row (§5.10)"),
    ]
    for i, (fill, label) in enumerate(legend_items):
        ws.cell(row=leg_row + i + 1, column=1).fill = fill
        ws.cell(row=leg_row + i + 1, column=1).border = BOX
        ws.cell(row=leg_row + i + 1, column=2, value=label)

    ws._sop_flags = flags


# ── Tab 2: Payroll Hours ─────────────────────────────────────────────────────

def _sheet_payroll_hours(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Payroll Hours")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="Payroll Hours Spreadsheet  (§5.11)")
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = DARK_BLUE
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    n = ws.cell(row=2, column=1,
                value="Regular ≤ 40 hrs  |  OT = hours > 40  |  PTO excluded from OT "
                      "threshold  |  ⚠ Holiday & Sick/Flex = fill in manually")
    n.font = Font(italic=True, size=9, color="FFFFFF")
    n.fill = MED_BLUE
    n.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    hdrs   = ["Last Name", "First Name", "Regular", "Vacation",
              "Holiday",  "Sick/Flex", "Overtime", "Total"]
    widths = [18, 16, 12, 12, 12, 12, 12, 14]
    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        _c(ws, 3, c, h, fill=DARK_BLUE, font=WHITE_FONT, align="center")
    ws.row_dimensions[3].height = 26
    ws.freeze_panes = "A4"

    MANUAL_FILL = PatternFill("solid", fgColor="FFFDE7")   # light yellow = manual entry

    for r_idx, row in enumerate(rows, 4):
        fill = ALT_FILL if r_idx % 2 == 0 else WHITE_FILL
        _c(ws, r_idx, 1, row["Last Name"],  fill=fill, bold=True)
        _c(ws, r_idx, 2, row["First Name"], fill=fill)
        for c_idx, key in enumerate(["Regular","Vacation","Holiday",
                                     "Sick/Flex","Overtime","Total"], 3):
            _c(ws, r_idx, c_idx, row[key], fill=fill, align="center", fmt="0.00")
        if row["_ot_flag"]:
            ws.cell(row=r_idx, column=7).fill = AMBER_FILL
            ws.cell(row=r_idx, column=7).font = BOLD_FONT
        if row["_pto_flag"]:
            ws.cell(row=r_idx, column=4).fill = LIGHT_BLUE
        # Holiday + Sick/Flex: light yellow to signal manual entry needed
        for col_idx in [5, 6]:
            ws.cell(row=r_idx, column=col_idx).fill = MANUAL_FILL

    # Totals row
    tot = len(rows) + 4
    ws.merge_cells(f"A{tot}:B{tot}")
    _c(ws, tot, 1, "TOTAL", fill=DARK_BLUE, font=WHITE_FONT, bold=True, align="center")
    ws.cell(row=tot, column=1).fill = DARK_BLUE
    ws.cell(row=tot, column=2).fill = DARK_BLUE
    ws.cell(row=tot, column=2).border = BOX
    for c in range(3, 9):
        ltr  = get_column_letter(c)
        cell = ws.cell(row=tot, column=c, value=f"=SUM({ltr}4:{ltr}{tot-1})")
        cell.fill = DARK_BLUE; cell.font = WHITE_FONT
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


# ── Tab 3: Flag Summary by Tech ──────────────────────────────────────────────

# (label, key-extractor-callable)  — key returns True when that flag is tripped
_FLAG_SPEC: list[tuple[str, callable]] = [
    ("OVER (p90)",    lambda r: bool(r.get("_thresh_result")) and r["_thresh_result"][0] == "OVER"),
    ("UNDER (p10)",   lambda r: bool(r.get("_thresh_result")) and r["_thresh_result"][0] == "UNDER"),
    ("No TSheets",    lambda r: not r.get("_tsheets_matched", True)),
    ("No Meld",       lambda r: r.get("_no_meld_entry", False)),
    ("NB No Note",    lambda r: r.get("_nb_missing_note", False)),
    ("Protected NB",  lambda r: r.get("_nb_overridden", False)),
    ("Pay>Bill",      lambda r: r["Paying"] > r["Billable"] and r["Billable"] > 0),
    ("Possible Dup",  lambda r: r.get("_possible_dup", False)),
    ("Overlap",       lambda r: r.get("_overlap_flag", False)),
    ("<0.10h",        lambda r: r.get("_under_01", False)),
    ("RC",            lambda r: r.get("_rc", False)),
    ("PS: Note",      lambda r: r.get("_has_ps_note", False)),
    ("15 Circle St",  lambda r: r.get("_is_15_circle", False)),
    ("Turnover",      lambda r: r.get("_turnover", False)),
]


def _sheet_flag_summary(wb: Workbook, recap_rows: list[dict]):
    """
    Per-tech rollup of flag counts (§5.9 / Q5).  One row per technician,
    one column per flag type plus Approved / Review counts and total hours.
    """
    ws = wb.create_sheet("Flag Summary by Tech")
    ws.sheet_view.showGridLines = False

    flag_labels = [s[0] for s in _FLAG_SPEC]
    hdrs   = ["Technician", "Total Entries", "✓ Approved", "⚠ Review",
              "Flagged Hours"] + flag_labels
    widths = [22, 12, 12, 12, 14] + [13] * len(flag_labels)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(hdrs))
    t = ws.cell(row=1, column=1,
                value="Flag Summary by Technician  (Per-Tech Rollup)")
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = DARK_BLUE
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(hdrs))
    n = ws.cell(row=2, column=1,
                value="Counts of each §5.9 flag per technician.  "
                      "Flagged Hours = sum of Actual hours on rows where Status = ⚠ Review.")
    n.font = Font(italic=True, size=9, color="FFFFFF")
    n.fill = MED_BLUE
    n.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    # Header row
    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        _c(ws, 3, c, h, fill=DARK_BLUE, font=WHITE_FONT, align="center", wrap=True)
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "B4"

    # Aggregate per tech (exclude non-labor admin/contact rows)
    techs: dict[str, list[dict]] = defaultdict(list)
    for r in recap_rows:
        if r.get("_non_labor"):
            continue
        techs[r["Tech"] or "(unknown)"].append(r)

    sorted_techs = sorted(techs.keys(), key=lambda n: n.lower())

    r_idx = 4
    for tech in sorted_techs:
        tech_rows = techs[tech]
        total_entries  = len(tech_rows)
        approved_ct    = sum(1 for r in tech_rows if r.get("_approval") == "Approved")
        review_ct      = total_entries - approved_ct
        flagged_hours  = sum(r["Actual"] for r in tech_rows
                             if r.get("_approval") == "Review Required")
        flag_counts    = [sum(1 for r in tech_rows if fn(r))
                          for _, fn in _FLAG_SPEC]

        base = ALT_FILL if r_idx % 2 == 0 else WHITE_FILL
        _c(ws, r_idx, 1, tech, fill=base, bold=True)
        _c(ws, r_idx, 2, total_entries, fill=base, align="center")
        _c(ws, r_idx, 3, approved_ct,
           fill=STATUS_APPROVED_FILL if approved_ct else base,
           align="center", bold=approved_ct > 0)
        _c(ws, r_idx, 4, review_ct,
           fill=STATUS_REVIEW_FILL if review_ct else base,
           align="center", bold=review_ct > 0)
        _c(ws, r_idx, 5, round(flagged_hours, 2),
           fill=base, align="center", fmt="0.00")
        for j, count in enumerate(flag_counts):
            cell_fill = base if count == 0 else AMBER_FILL
            _c(ws, r_idx, 6 + j, count if count else "",
               fill=cell_fill, align="center", bold=count > 0)
        r_idx += 1

    # Totals row
    if sorted_techs:
        tot_row = r_idx
        _c(ws, tot_row, 1, "TOTAL", fill=DARK_BLUE, font=WHITE_FONT,
           bold=True, align="center")
        for c in range(2, len(hdrs) + 1):
            ltr  = get_column_letter(c)
            cell = ws.cell(row=tot_row, column=c,
                           value=f"=SUM({ltr}4:{ltr}{tot_row - 1})")
            cell.fill = DARK_BLUE; cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BOX
            if c == 5:
                cell.number_format = "0.00"


# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Payroll Reconciliation – Stonelink",
        page_icon="📊",
        layout="wide",
    )

    st.title("📊 Weekly Payroll Reconciliation")
    st.caption(
        "Stonelink Property Management  ·  TSheets × Property Meld  ·  SOP v2.0 "
        "(Updated 4-16-2026)  ·  46-Category Task Thresholds + Approval Status Active"
    )

    c1, c2 = st.columns(2)
    with c1:
        pdf_file = st.file_uploader(
            "TSheets / QB Time Payroll PDF", type=["pdf"],
            help="Weekly payroll report — hours only source (§4.1)"
        )
    with c2:
        csv_file = st.file_uploader(
            "Property Meld Work Log Summary CSV", type=["csv"],
            help="PRIMARY source — must have all 7 required columns (§4.3)"
        )

    if not pdf_file or not csv_file:
        st.info("Upload both files above to generate the payroll report.")
        with st.expander("SOP v2.0 Quick Reference (Updated 4-16-2026)"):
            st.markdown("""
**§4.3 Required CSV Columns (7):**
Agent · Meld · Unit · Title · Description · Check-In Hours · Address Line 1

**§5.6 Billing Rules:**
| Type | Action |
|------|--------|
| Default | Billable unless explicitly changed |
| RC – Resident Charge | Note "RC – [reason]", **teal** highlight |
| Non-Billable | Note "Not billable – [reason]", H = 0 |
| Diagnostic / Inspection / Mitigation / Quote | **Always billable** — NB blocked |
| Flat bill | Exact phrase: `Not billable- to be flat bill` |

**§5.8 Approved Notes (5 only):**
`Not billable – [reason]` · `Combined above` · `[Initials] Hours combined`
`Adjusted for time` · `PS: [Instructions]` ← *PS: rows = yellow highlight*

**§5.9 Threshold Engine:** 46 task categories from 18,489 historical jobs.
Flags UNDER (< p10) and OVER (≥ p90) for each category.
**Turnover rows are skipped** (handled by turnover-specific logic).

**Approval Status (Col I):** Default = ✓ Approved.  ANY tripped flag overrides
the default and sets the row to **⚠ Review Required** (manual approval required).
Flat-bill entries are pre-approved and do NOT trip review.

**§5.11 Payroll Hours:** Last/First Name · Regular · Vacation (TSheets PTO)
· Holiday *(manual)* · Sick/Flex *(manual)* · Overtime · Total

**Excel Tab 3 — Flag Summary by Tech:** Per-tech rollup of flag counts
and approval-status totals for quick week-over-week pattern review.
            """)
        return

    with st.spinner("Parsing TSheets PDF…"):
        try:
            employees = parse_qb_pdf(pdf_file)
        except Exception as e:
            st.error(f"PDF parse error: {e}")
            return

    with st.expander("🔍 Debug – Raw PDF & Parsed Summaries", expanded=False):
        pdf_file.seek(0)
        import pdfplumber as _pl
        with _pl.open(pdf_file) as _pdf:
            _raw = "\n".join(p.extract_text() or "" for p in _pdf.pages)
        st.text_area("Raw PDF (first 4000 chars)", _raw[:4000], height=280)
        st.markdown("**Parsed employee summaries:**")
        for e in employees:
            st.json({k: v for k, v in e.items() if k != "shifts"})
        pdf_file.seek(0)

    with st.spinner("Loading Meld CSV…"):
        try:
            melds_df, csv_warnings = load_melds(csv_file)
        except Exception as e:
            st.error(f"CSV load error: {e}")
            return

    for w in csv_warnings:
        st.warning(w)

    with st.spinner("Applying SOP v2.0 rules + threshold checks…"):
        recap_rows   = build_recap_rows(employees, melds_df)
        payroll_rows = build_payroll_rows(employees, recap_rows)

    # Metrics
    st.divider()
    # Property Meld labor = CSV rows with hours, excluding No-Meld and non-labor.
    pm_rows        = [r for r in recap_rows
                      if not r.get("_no_meld_entry") and not r.get("_non_labor")]
    nomeld_rows    = [r for r in recap_rows if r.get("_no_meld_entry")]
    nonlabor_rows  = [r for r in recap_rows if r.get("_non_labor")]
    pm_actual      = sum(r["Actual"]   for r in pm_rows)
    pm_billable    = sum(r["Billable"] for r in pm_rows)
    nomeld_hrs     = sum(r["Actual"]   for r in nomeld_rows)

    no_meld_ct     = len(nomeld_rows)
    fifteen_ct     = sum(1 for r in recap_rows if r.get("_is_15_circle"))
    over_ct        = sum(1 for r in recap_rows
                         if r.get("_thresh_result") and r["_thresh_result"][0] == "OVER")
    under_ct       = sum(1 for r in recap_rows
                         if r.get("_thresh_result") and r["_thresh_result"][0] == "UNDER")
    dup_ct         = sum(1 for r in recap_rows if r.get("_possible_dup"))
    review_ct      = sum(1 for r in recap_rows if r.get("_approval") == "Review Required")
    approved_ct    = len(recap_rows) - review_ct
    turnover_ct    = sum(1 for r in recap_rows if r.get("_turnover"))
    total_ot       = sum(r["Overtime"] for r in payroll_rows)

    c = st.columns(9)
    c[0].metric("Employees",            len(employees))
    c[1].metric("Labor Entries",        len(pm_rows))
    c[2].metric("Meld Actual Hrs",      f"{pm_actual:.2f}")
    c[3].metric("Meld Billable Hrs",    f"{pm_billable:.2f}")
    c[4].metric("⚠ Review Required",    review_ct)
    c[5].metric("🔴 No-Meld Hrs",       f"{nomeld_hrs:.2f}")
    c[6].metric("⬆ Over Threshold",    over_ct)
    c[7].metric("⬇ Under Threshold",   under_ct)
    c[8].metric("OT Hrs",               f"{total_ot:.2f}")

    st.caption(
        f"**Meld Actual Hrs** = sum of Property Meld CSV Hours (the billing basis).  "
        f"**No-Meld Hrs** = {nomeld_hrs:.2f}h of TSheets time with no matching work order "
        f"(escalate, §7.3).  {len(nonlabor_rows)} zero-hour admin/contact entries are "
        f"shown in a separate section (not billed)."
    )

    if turnover_ct:
        st.info(f"🔁 {turnover_ct} **Turnover** rows detected — threshold checks "
                f"skipped per methodology (handled by turnover-specific logic).")

    # Actionable errors
    actionable = [
        f"**{r['Tech']}** – {r['_date']} – {r['MWO'] or '(no meld)'}: "
        f"Non-billable has NO explanation note (§5.5 REQUIRED)"
        for r in recap_rows if r.get("_nb_missing_note")
    ]
    if actionable:
        with st.expander(f"🚨 {len(actionable)} ACTIONABLE ERRORS – NB Notes Missing",
                         expanded=True):
            st.error("§5.5: A note is required for EVERY non-billable entry.")
            for e in actionable:
                st.error(e)

    violations = []
    for r in recap_rows:
        if r["Paying"] > r["Billable"] and r["Billable"] > 0:
            violations.append(
                f"**{r['Tech']}** – {r['_date']} – {r['MWO']}: "
                f"Paying {r['Paying']:.2f} > Billable {r['Billable']:.2f}"
            )
    if violations:
        with st.expander(f"⚠ {len(violations)} Paying > Billable (§6)", expanded=False):
            for v in violations:
                st.warning(v)

    # Tabs
    t1, t2, t3 = st.tabs(["Weekly Recap", "Payroll Hours", "Flags & Exceptions"])

    with t1:
        dc = ["Tech","_date","MWO","Address","Unit","Check-In",
              "Actual","Paying","Billable","_approval","Notes","Trade"]
        labor_display = [r for r in recap_rows if not r.get("_non_labor")]
        df_d = (pd.DataFrame(labor_display)[dc]
                .rename(columns={"_date": "Date", "_approval": "Status"}))
        st.dataframe(df_d, use_container_width=True, hide_index=True,
                     column_config={
                         "Actual":   st.column_config.NumberColumn(format="%.2f"),
                         "Paying":   st.column_config.NumberColumn(format="%.2f"),
                         "Billable": st.column_config.NumberColumn(format="%.2f"),
                     })
        if fifteen_ct:
            st.info(f"🟣 {fifteen_ct} entries at 15 Circle St are included above "
                    f"but appear in a SEPARATE section in the Excel output.")
        if nonlabor_rows:
            with st.expander(f"📋 {len(nonlabor_rows)} Non-Labor / Admin Entries "
                             f"(resident contact, scoping, status — 0 hrs, not billed)"):
                df_nl = (pd.DataFrame(nonlabor_rows)[["Tech","_date","MWO","Address",
                                                      "Notes","Trade"]]
                         .rename(columns={"_date": "Date"}))
                st.dataframe(df_nl, use_container_width=True, hide_index=True)
        st.caption("Col G (Paying) = Actual (Property Meld CSV Hours) by default. "
                   "Adjust in Excel for unreasonable durations — manager approval + "
                   "Notes entry required (§5.5).")

    with t2:
        df_p = pd.DataFrame(payroll_rows).drop(columns=["_ot_flag","_pto_flag"])
        st.dataframe(df_p, use_container_width=True, hide_index=True,
                     column_config={
                         k: st.column_config.NumberColumn(format="%.2f")
                         for k in ["Regular","Vacation","Holiday","Sick/Flex","Overtime","Total"]
                     })
        st.caption(
            "Vacation = TSheets PTO total.  "
            "**Holiday and Sick/Flex are 0 — fill in manually** in the downloaded Excel.  "
            "PTO does NOT count toward 40-hr OT threshold (§5.11)."
        )

    with t3:
        # Per-tech flag rollup (Q5)
        st.subheader("📊 Flag Summary by Technician")
        st.caption("One row per tech.  Counts of each §5.9 flag and approval-status totals.")

        techs: dict[str, list[dict]] = defaultdict(list)
        for r in recap_rows:
            if r.get("_non_labor"):
                continue
            techs[r["Tech"] or "(unknown)"].append(r)

        rollup_rows = []
        for tech in sorted(techs.keys(), key=lambda n: n.lower()):
            tr = techs[tech]
            total_entries = len(tr)
            appr = sum(1 for r in tr if r.get("_approval") == "Approved")
            rev  = total_entries - appr
            flagged_hours = sum(r["Actual"] for r in tr
                                if r.get("_approval") == "Review Required")
            row = {
                "Tech":            tech,
                "Entries":         total_entries,
                "✓ Approved":      appr,
                "⚠ Review":        rev,
                "Flagged Hrs":     round(flagged_hours, 2),
            }
            for label, fn in _FLAG_SPEC:
                row[label] = sum(1 for r in tr if fn(r))
            rollup_rows.append(row)

        if rollup_rows:
            st.dataframe(pd.DataFrame(rollup_rows),
                         use_container_width=True, hide_index=True)
        st.divider()

        # §5.9 — 15 Circle St
        c15 = [r for r in recap_rows if r.get("_is_15_circle")]
        if c15:
            st.error(f"🟣 {len(c15)} entries at **15 Circle St** — "
                     f"PROCESS SEPARATELY, do NOT include in standard billing (§5.9/§6)")
            st.dataframe(
                pd.DataFrame(c15)[["Tech","_date","MWO","Actual","Trade","Address"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)
        else:
            st.success("✅ No 15 Circle St entries.")

        # §7.3 — No Meld (unmatched TSheets)
        nm = [r for r in recap_rows if r.get("_no_meld_entry")]
        if nm:
            st.error(f"🔴 {len(nm)} TSheets shifts have **no Property Meld work order** — "
                     f"escalate to manager (§7.3)")
            st.dataframe(
                pd.DataFrame(nm)[["Tech","_date","Actual","_raw_notes"]]
                .rename(columns={"_date":"Date","_raw_notes":"TSheets Notes"}),
                use_container_width=True, hide_index=True)

        # §5.9 — Threshold OVER
        ov = [r for r in recap_rows
              if r.get("_thresh_result") and r["_thresh_result"][0] == "OVER"]
        if ov:
            st.warning(f"⬆ {len(ov)} entries **exceed p90 threshold** — "
                       f"review with manager (§5.9)")
            st.dataframe(pd.DataFrame([{
                "Tech": r["Tech"], "Date": r["_date"], "MWO": r["MWO"],
                "Actual": r["Actual"], "Trade": r["Trade"],
                "Detail": r["_thresh_result"][1]
            } for r in ov]), use_container_width=True, hide_index=True)
        else:
            st.success("✅ No entries exceed p90 task threshold.")

        # §5.9 — Threshold UNDER
        un = [r for r in recap_rows
              if r.get("_thresh_result") and r["_thresh_result"][0] == "UNDER"]
        if un:
            st.warning(f"⬇ {len(un)} entries **below p10 threshold** — "
                       f"possible missing time (§5.9)")
            st.dataframe(pd.DataFrame([{
                "Tech": r["Tech"], "Date": r["_date"], "MWO": r["MWO"],
                "Actual": r["Actual"], "Trade": r["Trade"],
                "Detail": r["_thresh_result"][1]
            } for r in un]), use_container_width=True, hide_index=True)
        else:
            st.success("✅ All entries within normal threshold band.")

        # §5.9 — Possible Duplicates
        dp = [r for r in recap_rows if r.get("_possible_dup")]
        if dp:
            st.warning(f"🔁 {len(dp)} rows flagged as **Possible Duplicates** "
                       f"(same MWO + Tech + Date) — review and consolidate (§5.9)")
            st.dataframe(
                pd.DataFrame(dp)[["Tech","_date","MWO","Actual","Notes"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)
        else:
            st.success("✅ No possible duplicates detected.")

        # §5.9 — Overlapping Shifts
        ov2 = [r for r in recap_rows if r.get("_overlap_flag")]
        if ov2:
            st.warning(f"⏱ {len(ov2)} entries with **overlapping check-in times** "
                       f"for the same tech — cannot bill same time twice (§5.9)")
            st.dataframe(
                pd.DataFrame(ov2)[["Tech","_date","MWO","Check-In","Actual"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)
        else:
            st.success("✅ No overlapping shift conflicts.")

        # §5.9 — Under 0.10 hrs
        tiny = [r for r in recap_rows if r.get("_under_01")]
        if tiny:
            st.warning(f"⚠ {len(tiny)} entries with **actual hours < 0.10** — "
                       f"confirm with manager (§5.9)")
            st.dataframe(
                pd.DataFrame(tiny)[["Tech","_date","MWO","Actual","Trade"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)
        else:
            st.success("✅ No sub-0.10 hour entries.")

        # §5.6 — RC entries
        rc_rows = [r for r in recap_rows if r.get("_rc")]
        if rc_rows:
            st.info(f"🔵 {len(rc_rows)} **RC (Resident Charge)** entries (§5.6)")
            st.dataframe(
                pd.DataFrame(rc_rows)[["Tech","_date","MWO","Actual","Notes"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)

        # Turnover rows (skipped from threshold per methodology)
        turn = [r for r in recap_rows if r.get("_turnover")]
        if turn:
            st.info(f"🔁 {len(turn)} **Turnover** rows — threshold checks skipped "
                    f"(handled by turnover-specific logic)")
            st.dataframe(
                pd.DataFrame(turn)[["Tech","_date","MWO","Actual","Trade","Notes"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)

        # §5.5 — NB missing note
        nb_nn = [r for r in recap_rows if r.get("_nb_missing_note")]
        if nb_nn:
            st.error(f"🚨 {len(nb_nn)} non-billable entries with **no explanation note** "
                     f"(§5.5 HARD REQUIREMENT)")
            st.dataframe(
                pd.DataFrame(nb_nn)[["Tech","_date","MWO","Actual","_raw_notes"]]
                .rename(columns={"_date":"Date","_raw_notes":"TSheets Notes"}),
                use_container_width=True, hide_index=True)
        else:
            st.success("✅ All non-billable entries have explanation notes.")

        # §5.8 — PS: notes
        ps = [r for r in recap_rows if r.get("_has_ps_note")]
        if ps:
            st.warning(f"🟡 {len(ps)} entries with **PS: notes** — "
                       f"highlighted yellow in Excel for Planet Synergy team (§5.8)")
            st.dataframe(
                pd.DataFrame(ps)[["Tech","_date","MWO","Actual","Notes"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)

        # §5.6 — Protected trade NB override blocked
        prot = [r for r in recap_rows if r.get("_nb_overridden")]
        if prot:
            st.error(f"⛔ {len(prot)} entries: **NB override blocked** — "
                     f"Diagnostic/Inspection/Mitigation/Quote requires manager auth (§5.6)")
            st.dataframe(
                pd.DataFrame(prot)[["Tech","_date","MWO","Actual","Trade"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)

        # §5.5 — Flat-bill entries
        flat = [r for r in recap_rows if r.get("_flat_bill")]
        if flat:
            st.subheader(f"Flat-Bill Entries ({len(flat)})")
            st.caption(f"Total: **{sum(r['Actual'] for r in flat):.2f}h**")
            st.dataframe(
                pd.DataFrame(flat)[["Tech","_date","MWO","Actual","Notes"]]
                .rename(columns={"_date":"Date"}),
                use_container_width=True, hide_index=True)

        # OT employees
        ot_list = [r for r in payroll_rows if r["_ot_flag"]]
        if ot_list:
            st.subheader("Employees with Overtime")
            st.dataframe(
                pd.DataFrame(ot_list)[["Last Name","First Name","Regular",
                                       "Overtime","Vacation","Total"]],
                use_container_width=True, hide_index=True)

    # Download
    st.divider()
    with st.spinner("Building Excel workbook…"):
        excel_bytes = build_excel(recap_rows, payroll_rows, employees)

    week = employees[0]["period_start"].replace("/", "-") if employees else ""
    st.download_button(
        label="⬇ Download Payroll Report (.xlsx)",
        data=excel_bytes,
        file_name=f"payroll_report_{week}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

    st.caption(
        "**Post-download:** Col G (Paying) = Actual by default — adjust for unreasonable "
        "durations with a Notes entry.  Fill **Holiday** and **Sick/Flex** columns manually "
        "in Payroll Hours tab.  Yellow rows = PS: notes for Planet Synergy.  "
        "Lavender = 15 Circle St (separate section in Excel).  "
        "**Col I (Status):** ⚠ Review entries need manual approval before payroll close.  "
        "All adjustments require §5.5 notes."
    )


if __name__ == "__main__":
    main()
